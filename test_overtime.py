from __future__ import annotations

import os
import sqlite3
import subprocess
import sys
import tempfile
import unittest
from unittest import mock
from io import BytesIO
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[1]


class OvertimeCalculationTests(unittest.TestCase):
    def test_active_batch_index_migrates_without_blocking_reimport_cycles(self):
        from overtime import ensure_overtime_schema

        connection = sqlite3.connect(":memory:")
        ensure_overtime_schema(connection)
        connection.execute("DROP INDEX idx_overtime_batch_active_period")
        connection.execute(
            "CREATE UNIQUE INDEX idx_overtime_batch_active_checksum "
            "ON overtime_import_batches(period, file_sha256, is_active)"
        )

        ensure_overtime_schema(connection)
        indexes = {
            row[0]
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'index'"
            )
        }
        self.assertNotIn("idx_overtime_batch_active_checksum", indexes)
        self.assertIn("idx_overtime_batch_active_period", indexes)

        values = ("2026-08", "sample.xls", "a" * 64, 1, "admin", "2026-08-17T00:00:00")
        connection.execute(
            "INSERT INTO overtime_import_batches "
            "(id, period, file_name, file_sha256, row_count, imported_by_username, imported_at, is_active) "
            "VALUES ('old-a', ?, ?, ?, ?, ?, ?, 0)",
            values,
        )
        connection.execute(
            "INSERT INTO overtime_import_batches "
            "(id, period, file_name, file_sha256, row_count, imported_by_username, imported_at, is_active) "
            "VALUES ('old-b', ?, ?, ?, ?, ?, ?, 0)",
            values,
        )
        connection.execute(
            "INSERT INTO overtime_import_batches "
            "(id, period, file_name, file_sha256, row_count, imported_by_username, imported_at, is_active) "
            "VALUES ('active-a', ?, ?, ?, ?, ?, ?, 1)",
            values,
        )
        with self.assertRaises(sqlite3.IntegrityError):
            connection.execute(
                "INSERT INTO overtime_import_batches "
                "(id, period, file_name, file_sha256, row_count, imported_by_username, imported_at, is_active) "
                "VALUES ('active-b', ?, ?, ?, ?, ?, ?, 1)",
                ("2026-08", "other.xls", "b" * 64, 1, "admin", "2026-08-17T00:01:00"),
            )
        connection.close()

    def test_contract_hour_formula(self):
        from overtime import calculate_live_hours

        self.assertEqual(calculate_live_hours(4, "workday"), 7.5)
        self.assertEqual(calculate_live_hours(4, "day_off"), 8.0)
        self.assertEqual(calculate_live_hours(0.5, "workday"), 0.75)
        self.assertEqual(calculate_live_hours(1.5, "workday"), 2.5)
        self.assertIsNone(calculate_live_hours(4, "unknown"))
        # With deduction
        self.assertEqual(calculate_live_hours(8, "day_off", deduction_hours=1.0), 14.0)
        self.assertEqual(calculate_live_hours(8, "workday", deduction_hours=1.0), 13.5)
        self.assertEqual(calculate_live_hours(4, "day_off", deduction_hours=1.0), 6.0)
        self.assertEqual(calculate_live_hours(4, "workday", deduction_hours=1.0), 5.5)

    def test_overtime_deduction_for_sbg_and_smor(self):
        from overtime import calculate_overtime_deduction

        # 1. PT SBG & PT SMOR on day_off -> deduction 1.0
        deduction, key, reason = calculate_overtime_deduction("sbg", "day_off", 8.0)
        self.assertEqual(deduction, 1.0)
        self.assertEqual(key, "day_off")
        self.assertIn("Libur", reason)

        deduction, key, reason = calculate_overtime_deduction("smor", "day_off", 4.0)
        self.assertEqual(deduction, 1.0)
        self.assertEqual(key, "day_off")

        # 2. PT SBG & PT SMOR on workday continuing shift 1 -> 2 (end >= 22:30 or duration >= 7)
        deduction, key, reason = calculate_overtime_deduction(
            "smor", "workday", 8.0, start_time="15:00:00", end_time="23:00:00", schedule_shift_code="1"
        )
        self.assertEqual(deduction, 1.0)
        self.assertEqual(key, "shift_1_2")
        self.assertIn("Shift 1→2", reason)

        # 3. PT SBG on workday continuing shift 2 -> 3 (start ~23:00, end ~07:00)
        deduction, key, reason = calculate_overtime_deduction(
            "sbg", "workday", 8.0, start_time="23:00:00", end_time="07:00:00", schedule_shift_code="2"
        )
        self.assertEqual(deduction, 1.0)
        self.assertEqual(key, "shift_2_3")
        self.assertIn("Shift 2→3", reason)

        # 4. PT SMOR on regular overtime workday (e.g. 4 hours from 16:30 to 20:30) -> no deduction
        deduction, key, reason = calculate_overtime_deduction(
            "smor", "workday", 4.0, start_time="16:30:00", end_time="20:30:00", schedule_shift_code="1"
        )
        self.assertEqual(deduction, 0.0)
        self.assertEqual(key, "")

        # 5. PT SONAR on day_off or shift 1->2 -> no deduction (remains standard)
        deduction, key, reason = calculate_overtime_deduction("sonar", "day_off", 8.0)
        self.assertEqual(deduction, 0.0)

        deduction, key, reason = calculate_overtime_deduction(
            "sonar", "workday", 8.0, start_time="15:00:00", end_time="23:00:00", schedule_shift_code="1"
        )
        self.assertEqual(deduction, 0.0)

    def test_contracts_are_exposed_as_company_names_only(self):
        from overtime import company_from_contract

        self.assertEqual(
            company_from_contract("JASA TENAGA PEMELIHARAAN DAN AUTONOMOUS  PERALATAN PENUNJANG PABRIK TU"),
            ("smor", "PT SMOR"),
        )
        self.assertEqual(
            company_from_contract("NAMA KONTRAK JASA PENUNJANG PABRIK TUBAN"),
            ("sbg", "PT SBG"),
        )
        self.assertEqual(
            company_from_contract("JASA PEMELIHARAAN RUTIN LISTRIK & INSTRUMEN AREA UNIT MAINTENANCE II P"),
            ("sonar", "PT SONAR"),
        )
        self.assertEqual(company_from_contract("Kontrak lain"), ("unmapped", "Belum dipetakan"))

    def test_schedule_revision_becomes_source_of_truth_and_contract_has_start_date(self):
        from openpyxl import Workbook
        from overtime import _schedule_sheet_metadata, is_quota_eligible, parse_overtime_schedule_xlsx

        workbook = Workbook()
        base = workbook.active
        base.title = "008"
        revision = workbook.create_sheet("008R01")
        for sheet, first_day in ((base, "AD"), (revision, "OFF")):
            sheet.cell(3, 10, "Agustus 2026")
            sheet.cell(8, 5, 1)
            sheet.cell(8, 6, 2)
            sheet.cell(10, 2, "94S25279")
            sheet.cell(10, 3, "Faiz Zamani")
            sheet.cell(10, 5, first_day)
            sheet.cell(10, 6, "OFF")
        buffer = BytesIO()
        workbook.save(buffer)

        entries, sheets = parse_overtime_schedule_xlsx(buffer.getvalue())
        by_date = {entry["workDate"]: entry for entry in entries}
        self.assertEqual(by_date["2026-08-01"]["dayType"], "day_off")
        self.assertEqual(by_date["2026-08-01"]["sourceSheet"], "008R01")
        self.assertEqual(by_date["2026-08-02"]["dayType"], "day_off")
        self.assertEqual(next(sheet for sheet in sheets if sheet["name"] == "008R01")["revisionRank"], 1)
        self.assertEqual(_schedule_sheet_metadata("R01006"), (6, 1))
        self.assertFalse(is_quota_eligible("2026-08-01", "smor"))
        self.assertTrue(is_quota_eligible("2026-08-02", "smor"))

    def test_company_contract_periods_are_inclusive_and_independent(self):
        from overtime import contract_period_for_company, is_quota_eligible

        self.assertEqual(contract_period_for_company("sbg")["startDate"], "2026-03-01")
        self.assertEqual(contract_period_for_company("sbg")["endDate"], "2027-02-28")
        self.assertTrue(is_quota_eligible("2026-03-01", "sbg"))
        self.assertTrue(is_quota_eligible("2027-02-28", "sbg"))
        self.assertFalse(is_quota_eligible("2026-02-28", "sbg"))
        self.assertFalse(is_quota_eligible("2027-03-01", "sbg"))
        self.assertTrue(is_quota_eligible("2026-08-02", "smor"))
        self.assertTrue(is_quota_eligible("2027-08-01", "sonar"))
        self.assertFalse(is_quota_eligible("2026-08-01", "smor"))
        self.assertFalse(is_quota_eligible("2027-08-02", "sonar"))

    def test_schedule_calculates_live_hours_by_unique_name_when_badge_changes(self):
        from overtime import ensure_overtime_schema, list_overtime

        connection = sqlite3.connect(":memory:")
        connection.row_factory = sqlite3.Row
        ensure_overtime_schema(connection)
        connection.execute(
            "INSERT INTO overtime_schedule_batches (id, file_name, file_sha256, sheet_count, row_count, imported_by_username, imported_at, is_active) VALUES ('schedule-1', 'jadwal.xlsx', 'a', 1, 1, 'admin', '2026-08-20T00:00:00', 1)"
        )
        connection.execute(
            "INSERT INTO overtime_schedule_entries (batch_id, employee_no, employee_name, work_date, shift_code, day_type, source_sheet, revision_rank) VALUES ('schedule-1', '82F347', 'Mudi', '2026-08-17', 'AD', 'workday', '008', 0)"
        )
        connection.execute(
            "INSERT INTO overtime_import_batches (id, period, file_name, file_sha256, row_count, imported_by_username, imported_at, is_active) VALUES ('batch-1', '2026-08', 'lembur.xls', 'b', 1, 'admin', '2026-08-20T00:00:00', 1)"
        )
        connection.execute(
            "INSERT INTO overtime_entries (id, period, batch_id, source_row_no, employee_no, employee_name, work_date, raw_hours, source_work_group, source_fingerprint, created_at) VALUES ('entry-1', '2026-08', 'batch-1', 2, '82I22886', 'MUDI', '2026-08-17', 4.0, 'JASA PENUNJANG PABRIK TUBAN', 'fingerprint-1', '2026-08-20T00:00:00')"
        )
        payload = list_overtime(lambda: connection, {"period": "2026-08", "page": "1", "limit": "10"})
        item = payload["items"][0]
        self.assertEqual(item["classificationSource"], "schedule")
        self.assertEqual(item["scheduleMatch"], "name")
        self.assertEqual(item["dayType"], "workday")
        self.assertEqual(item["liveHours"], 7.5)
        connection.close()

    def test_multi_month_import_derives_periods_and_replaces_each_month_atomically(self):
        from overtime import ensure_overtime_schema, import_overtime_xls, preview_overtime_xls

        connection = sqlite3.connect(":memory:")
        connection.row_factory = sqlite3.Row
        ensure_overtime_schema(connection)
        rows = [
            {
                "id": "ot-mar", "period": "2026-03", "sourceRowNo": 2, "sourceNumber": "1", "sourceStatus": "APPROVED",
                "employeeNo": "EMP-1", "employeeName": "HASAN", "workDate": "2026-03-01", "startTime": "16:30:00",
                "endTime": "20:30:00", "task": "Test", "rawHours": 4.0, "clockDurationHours": 4.0,
                "durationDelta": 0.0, "unitName": "Listrik", "sourceWorkGroup": "JASA PENUNJANG PABRIK TUBAN",
                "approver": "Admin", "approvedAt": "", "sourceFingerprint": "fingerprint-mar",
            },
            {
                "id": "ot-aug", "period": "2026-08", "sourceRowNo": 3, "sourceNumber": "2", "sourceStatus": "PENDING",
                "employeeNo": "EMP-1", "employeeName": "HASAN", "workDate": "2026-08-17", "startTime": "07:30:00",
                "endTime": "11:30:00", "task": "Test", "rawHours": 4.0, "clockDurationHours": 4.0,
                "durationDelta": 0.0, "unitName": "Listrik", "sourceWorkGroup": "JASA PENUNJANG PABRIK TUBAN",
                "approver": "Admin", "approvedAt": "", "sourceFingerprint": "fingerprint-aug",
            },
        ]
        payload = {"fileName": "lembur.xls", "fileData": "ignored"}
        with mock.patch("overtime.decode_overtime_file", return_value=b"workbook"), mock.patch("overtime.parse_overtime_xls", return_value=rows):
            preview = preview_overtime_xls(payload, 1024, lambda: connection)
            self.assertEqual(preview["periods"], ["2026-03", "2026-08"])
            self.assertEqual(preview["latestPeriod"], "2026-08")
            result = import_overtime_xls(lambda: connection, payload, {"id": 1, "username": "admin"}, "2026-08-20T00:00:00", 1024)
            self.assertFalse(result["noOp"])
            self.assertEqual(result["periods"], ["2026-03", "2026-08"])
            self.assertEqual(result["imported"], 2)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM overtime_entries").fetchone()[0], 2)
            second = import_overtime_xls(lambda: connection, payload, {"id": 1, "username": "admin"}, "2026-08-20T00:01:00", 1024)
            self.assertTrue(second["noOp"])
        connection.close()

    def test_summary_classification_group_and_quota(self):
        script = r'''
import server
from overtime import list_overtime, update_overtime_day_types, update_overtime_personnel

server.init_db()
admin = server.get_user_by_username("admin.plirm34")
now = server.utc_now().isoformat()
with server.get_connection() as connection:
    connection.execute(
        "INSERT INTO overtime_import_batches (id, period, file_name, file_sha256, row_count, imported_by_user_id, imported_by_username, imported_at, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)",
        ("batch-1", "2026-07", "sample.xls", "a" * 64, 3, int(admin["id"]), admin["username"], now),
    )
    for employee_no, name in (("EMP-1", "HASAN"), ("EMP-2", "SITI")):
        connection.execute(
            "INSERT INTO overtime_personnel (employee_no, employee_name, group_type, updated_by_user_id, updated_at) VALUES (?, ?, 'unassigned', ?, ?)",
            (employee_no, name, int(admin["id"]), now),
        )
    entries = [
        ("e1", 2, "APPROVED", "EMP-1", "HASAN", "2026-07-01", 4.0, "f1"),
        ("e2", 3, "REJECTED", "EMP-1", "HASAN", "2026-07-02", 4.0, "f2"),
        ("e3", 4, "PENDING", "EMP-2", "SITI", "2026-07-03", 1.5, "f3"),
    ]
    for item_id, row_no, status, employee_no, name, work_date, raw_hours, fingerprint in entries:
        connection.execute(
            """
            INSERT INTO overtime_entries (
                id, period, batch_id, source_row_no, source_status, employee_no, employee_name,
                work_date, start_time, end_time, task, raw_hours, clock_duration_hours,
                duration_delta, source_work_group, source_fingerprint, created_at
            ) VALUES (?, '2026-07', 'batch-1', ?, ?, ?, ?, ?, '16:30:00', '20:30:00', 'Test', ?, 4.0, 0.0, 'JASA PENUNJANG PABRIK TUBAN', ?, ?)
            """,
            (item_id, row_no, status, employee_no, name, work_date, raw_hours, fingerprint, now),
        )

before = list_overtime(server.get_connection, {"period": "2026-07", "page": "1", "limit": "10"})
assert before["summary"]["unclassifiedCount"] == 3
assert before["summary"]["totalLiveHours"] == 0

update_overtime_day_types(server.get_connection, [
    {"employeeNo": "EMP-1", "workDate": "2026-07-01", "dayType": "workday"},
    {"employeeNo": "EMP-1", "workDate": "2026-07-02", "dayType": "day_off"},
    {"employeeNo": "EMP-2", "workDate": "2026-07-03", "dayType": "workday"},
], int(admin["id"]), now)
update_overtime_personnel(server.get_connection, "EMP-1", "gangguan", int(admin["id"]), now)
update_overtime_personnel(server.get_connection, "EMP-2", "preventif", int(admin["id"]), now)

after = list_overtime(server.get_connection, {"period": "2026-07", "page": "1", "limit": "10"})
assert after["summary"]["unclassifiedCount"] == 0
assert after["summary"]["totalLiveHours"] == 16.0
assert {item["status"] for item in after["items"]} == {"APPROVED", "REJECTED", "PENDING"}
hasan = next(item for item in after["summary"]["ranking"] if item["employeeNo"] == "EMP-1")
assert hasan["monthLiveHours"] == 13.5
assert hasan["contractLiveHours"] == 13.5
assert hasan["annualRemainingHours"] == 370.5
assert hasan["groupType"] == "gangguan"
groups = {item["groupType"]: item for item in after["summary"]["groupTotals"]}
assert groups["gangguan"]["liveHours"] == 13.5
assert groups["preventif"]["liveHours"] == 2.5

page = list_overtime(server.get_connection, {"period": "2026-07", "page": "2", "limit": "10", "status": "APPROVED"})
assert page["pagination"]["total"] == 1
assert page["pagination"]["pages"] == 1
print("overtime_summary=ok")
'''
        with tempfile.TemporaryDirectory() as temporary_directory:
            environment = os.environ.copy()
            environment["PLIRM34_DATA_DIR"] = temporary_directory
            environment["PLIRM34_SKIP_LEGACY_MIGRATION"] = "1"
            result = subprocess.run(
                [sys.executable, "-c", script],
                cwd=PROJECT_ROOT,
                env=environment,
                capture_output=True,
                text=True,
                timeout=45,
                check=False,
            )
        self.assertEqual(result.returncode, 0, msg=f"stdout:\n{result.stdout}\nstderr:\n{result.stderr}")

    def test_sbg_contract_summary_crosses_calendar_year_without_merging_pt(self):
        script = r'''
import server
from overtime import list_overtime, update_overtime_day_types

server.init_db()
admin = server.get_user_by_username("admin.plirm34")
now = server.utc_now().isoformat()
with server.get_connection() as connection:
    for period, batch in (("2026-03", "batch-mar"), ("2026-08", "batch-aug"), ("2027-01", "batch-jan")):
        connection.execute(
            "INSERT INTO overtime_import_batches (id, period, file_name, file_sha256, row_count, imported_by_username, imported_at, is_active) VALUES (?, ?, 'sample.xls', ?, 1, 'admin', ?, 1)",
            (batch, period, batch * 8, now),
        )
    connection.execute("INSERT INTO overtime_personnel (employee_no, employee_name, group_type, updated_at) VALUES ('EMP-1', 'HASAN', 'gangguan', ?)", (now,))
    rows = [
        ("sbg-mar", "2026-03", "batch-mar", "2026-03-01", 4.0, "JASA PENUNJANG PABRIK TUBAN"),
        ("smor-before", "2026-08", "batch-aug", "2026-08-01", 4.0, "JASA TENAGA PEMELIHARAAN DAN AUTONOMOUS PERALATAN PENUNJANG PABRIK TU"),
        ("smor-aug", "2026-08", "batch-aug", "2026-08-02", 4.0, "JASA TENAGA PEMELIHARAAN DAN AUTONOMOUS PERALATAN PENUNJANG PABRIK TU"),
        ("sbg-jan", "2027-01", "batch-jan", "2027-01-15", 4.0, "JASA PENUNJANG PABRIK TUBAN"),
    ]
    for item_id, period, batch, work_date, raw_hours, contract in rows:
        connection.execute(
            "INSERT INTO overtime_entries (id, period, batch_id, source_row_no, employee_no, employee_name, work_date, raw_hours, source_work_group, source_fingerprint, created_at) VALUES (?, ?, ?, 1, 'EMP-1', 'HASAN', ?, ?, ?, ?, ?)",
            (item_id, period, batch, work_date, raw_hours, contract, item_id, now),
        )

update_overtime_day_types(server.get_connection, [
    {"employeeNo": "EMP-1", "workDate": "2026-03-01", "dayType": "day_off"},
    {"employeeNo": "EMP-1", "workDate": "2026-08-01", "dayType": "day_off"},
    {"employeeNo": "EMP-1", "workDate": "2026-08-02", "dayType": "day_off"},
    {"employeeNo": "EMP-1", "workDate": "2027-01-15", "dayType": "day_off"},
], int(admin["id"]), now)
payload = list_overtime(server.get_connection, {"period": "2027-01", "page": "1", "limit": "10"})
ranking = payload["summary"]["ranking"]
assert len(ranking) == 1
assert ranking[0]["companyKey"] == "sbg"
assert ranking[0]["contractLiveHours"] == 12.0
assert ranking[0]["annualRemainingHours"] == 372.0
admin_items = list_overtime(server.get_connection, {"period": "2026-08", "page": "1", "limit": "10"})["summary"]["ranking"]
assert len(admin_items) == 1 and admin_items[0]["companyKey"] == "smor"
assert admin_items[0]["contractLiveHours"] == 6.0
august = list_overtime(server.get_connection, {"period": "2026-08", "page": "1", "limit": "10"})["items"]
outside = next(item for item in august if item["id"] == "smor-before")
assert outside["quotaEligible"] is False and outside["quotaEligibilityLabel"] == "Di luar periode kontrak"
smor_range = list_overtime(server.get_connection, {
    "startDate": "2026-08-01", "endDate": "2026-10-31", "company": "smor", "page": "1", "limit": "10",
})
assert smor_range["range"] == {"startDate": "2026-08-01", "endDate": "2026-10-31"}
assert len(smor_range["items"]) == 2
assert smor_range["summary"]["ranking"][0]["companyKey"] == "smor"
print("cross_year_company_quota=ok")
'''
        with tempfile.TemporaryDirectory() as temporary_directory:
            environment = os.environ.copy()
            environment["PLIRM34_DATA_DIR"] = temporary_directory
            environment["PLIRM34_SKIP_LEGACY_MIGRATION"] = "1"
            result = subprocess.run(
                [sys.executable, "-c", script], cwd=PROJECT_ROOT, env=environment,
                capture_output=True, text=True, timeout=45, check=False,
            )
        self.assertEqual(result.returncode, 0, msg=f"stdout:\n{result.stdout}\nstderr:\n{result.stderr}")


if __name__ == "__main__":
    unittest.main()

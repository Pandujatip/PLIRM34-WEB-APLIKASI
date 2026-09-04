from __future__ import annotations

import base64
import binascii
import hashlib
import io
import math
import re
import sqlite3
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Callable


OVERTIME_ANNUAL_QUOTA = 384.0
OVERTIME_MONTHLY_GUIDELINE = 32.0
OVERTIME_WARNING_RATIO = 0.8
OVERTIME_GROUPS = {"unassigned", "gangguan", "preventif"}
OVERTIME_DAY_TYPES = {"unknown", "workday", "day_off"}
OVERTIME_COMPANIES = {
    "smor": "PT SMOR",
    "sbg": "PT SBG",
    "sonar": "PT SONAR",
    "unmapped": "Belum dipetakan",
}
OVERTIME_COMPANY_CONTRACTS = (
    ("JASA TENAGA PEMELIHARAAN DAN AUTONOMOUS PERALATAN PENUNJANG PABRIK TU", "smor"),
    ("JASA PENUNJANG PABRIK TUBAN", "sbg"),
    ("JASA PEMELIHARAAN RUTIN LISTRIK & INSTRUMEN AREA UNIT MAINTENANCE II P", "sonar"),
)
OVERTIME_COMPANY_CONTRACT_PERIODS = {
    "smor": {"startDate": "2026-08-02", "endDate": "2027-08-01"},
    "sbg": {"startDate": "2026-03-01", "endDate": "2027-02-28"},
    "sonar": {"startDate": "2026-08-02", "endDate": "2027-08-01"},
}
OVERTIME_REQUIRED_HEADERS = {
    "no",
    "status",
    "no pegawai",
    "nama pegawai",
    "tanggal",
    "jam mulai",
    "jam selesai",
    "tugas",
    "total",
    "unit kerja",
    "kelompok kerja",
    "app by",
    "app datetime",
}
OVERTIME_SCHEDULE_OFF_CODES = {"OFF", "0FF", "LIBUR", "HB", "HB/CT"}


def ensure_overtime_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS overtime_import_batches (
            id TEXT PRIMARY KEY,
            period TEXT NOT NULL,
            file_name TEXT NOT NULL,
            file_sha256 TEXT NOT NULL,
            row_count INTEGER NOT NULL DEFAULT 0,
            imported_by_user_id INTEGER,
            imported_by_username TEXT NOT NULL DEFAULT '',
            imported_at TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            replaced_at TEXT NOT NULL DEFAULT ''
        );

        CREATE UNIQUE INDEX IF NOT EXISTS idx_overtime_batch_active_period
        ON overtime_import_batches(period)
        WHERE is_active = 1;

        CREATE TABLE IF NOT EXISTS overtime_personnel (
            employee_no TEXT PRIMARY KEY,
            employee_name TEXT NOT NULL,
            group_type TEXT NOT NULL DEFAULT 'unassigned'
                CHECK (group_type IN ('unassigned', 'gangguan', 'preventif')),
            updated_by_user_id INTEGER,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS overtime_day_types (
            employee_no TEXT NOT NULL,
            work_date TEXT NOT NULL,
            day_type TEXT NOT NULL DEFAULT 'unknown'
                CHECK (day_type IN ('unknown', 'workday', 'day_off')),
            updated_by_user_id INTEGER,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (employee_no, work_date)
        );

        CREATE TABLE IF NOT EXISTS overtime_entries (
            id TEXT PRIMARY KEY,
            period TEXT NOT NULL,
            batch_id TEXT NOT NULL,
            source_row_no INTEGER NOT NULL,
            source_number TEXT NOT NULL DEFAULT '',
            source_status TEXT NOT NULL DEFAULT '',
            employee_no TEXT NOT NULL,
            employee_name TEXT NOT NULL,
            work_date TEXT NOT NULL,
            start_time TEXT NOT NULL DEFAULT '',
            end_time TEXT NOT NULL DEFAULT '',
            task TEXT NOT NULL DEFAULT '',
            raw_hours REAL NOT NULL,
            clock_duration_hours REAL,
            duration_delta REAL,
            unit_name TEXT NOT NULL DEFAULT '',
            source_work_group TEXT NOT NULL DEFAULT '',
            approver TEXT NOT NULL DEFAULT '',
            approved_at TEXT NOT NULL DEFAULT '',
            source_fingerprint TEXT NOT NULL,
            created_at TEXT NOT NULL,
            UNIQUE (period, source_fingerprint)
        );

        CREATE INDEX IF NOT EXISTS idx_overtime_entries_period
        ON overtime_entries(period, work_date, employee_no);

        CREATE INDEX IF NOT EXISTS idx_overtime_entries_employee
        ON overtime_entries(employee_no, work_date);

        CREATE TABLE IF NOT EXISTS overtime_schedule_batches (
            id TEXT PRIMARY KEY,
            file_name TEXT NOT NULL,
            file_sha256 TEXT NOT NULL,
            sheet_count INTEGER NOT NULL DEFAULT 0,
            row_count INTEGER NOT NULL DEFAULT 0,
            imported_by_user_id INTEGER,
            imported_by_username TEXT NOT NULL DEFAULT '',
            imported_at TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            replaced_at TEXT NOT NULL DEFAULT ''
        );

        CREATE UNIQUE INDEX IF NOT EXISTS idx_overtime_schedule_active
        ON overtime_schedule_batches(is_active)
        WHERE is_active = 1;

        CREATE TABLE IF NOT EXISTS overtime_schedule_entries (
            batch_id TEXT NOT NULL,
            employee_no TEXT NOT NULL,
            employee_name TEXT NOT NULL DEFAULT '',
            work_date TEXT NOT NULL,
            shift_code TEXT NOT NULL DEFAULT '',
            day_type TEXT NOT NULL CHECK (day_type IN ('workday', 'day_off')),
            source_sheet TEXT NOT NULL,
            revision_rank INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (batch_id, employee_no, work_date)
        );

        CREATE INDEX IF NOT EXISTS idx_overtime_schedule_lookup
        ON overtime_schedule_entries(batch_id, employee_no, work_date);
        """
    )
    legacy_index = connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'index' AND name = 'idx_overtime_batch_active_checksum'"
    ).fetchone()
    if legacy_index:
        connection.execute("DROP INDEX idx_overtime_batch_active_checksum")


OVERTIME_DEDUCTION_COMPANIES = {"sbg", "smor"}


def calculate_overtime_deduction(
    company_key: str,
    day_type: str,
    raw_hours: float,
    start_time: str = "",
    end_time: str = "",
    schedule_shift_code: str = "",
) -> tuple[float, str, str]:
    """
    Menghitung pemotongan 1 jam lembur untuk PT SBG & PT SMOR:
    - Hanya dipotong 1 jam jika jam lembur (raw_hours) >= 8.0 jam:
      1. Lembur pada hari libur (day_off) >= 8 jam
      2. Lembur nerus dari Shift 1 sampai akhir Shift 2 >= 8 jam
      3. Lembur nerus dari Shift 2 sampai akhir Shift 3 >= 8 jam
    - Jika raw_hours < 8.0 jam (misal 7 jam, atau 16:30-23:30), tidak dipotong lagi.

    Returns: (deduction_hours: float, deduction_key: str, deduction_reason: str)
    """
    if company_key not in OVERTIME_DEDUCTION_COMPANIES:
        return 0.0, "", ""

    raw = max(0.0, float(raw_hours or 0.0))
    if raw < 8.0:
        return 0.0, "", ""

    if day_type == "day_off":
        return 1.0, "day_off", "Lembur Hari Libur ≥ 8 jam (-1 jam)"

    if day_type == "workday":
        shift_code = re.sub(r"\s+", "", str(schedule_shift_code or "").upper())
        start = str(start_time or "").strip()
        end = str(end_time or "").strip()

        is_shift_1 = bool(shift_code in {"1", "S1", "P", "PAGI", "01", "N1", "AD", "NS"} or shift_code.startswith("1") or shift_code.startswith("P"))
        reaches_end_shift_2 = bool(end >= "22:30:00" or end >= "22:00:00")
        starts_after_shift_1 = bool(start >= "14:30:00" and start <= "18:00:00")

        if (is_shift_1 and reaches_end_shift_2) or (starts_after_shift_1 and reaches_end_shift_2) or reaches_end_shift_2:
            return 1.0, "shift_1_2", "Lembur Nerus Shift 1→2 ≥ 8 jam (-1 jam)"

        is_shift_2 = bool(shift_code in {"2", "S2", "S", "SIANG", "02", "N2"} or shift_code.startswith("2") or shift_code.startswith("S"))
        reaches_end_shift_3 = bool(end <= "08:30:00" and (start >= "22:00:00" or raw >= 7.5))
        starts_after_shift_2 = bool(start >= "22:00:00" or start >= "22:30:00")

        if (is_shift_2 and reaches_end_shift_3) or (starts_after_shift_2 and reaches_end_shift_3) or starts_after_shift_2 or reaches_end_shift_3:
            return 1.0, "shift_2_3", "Lembur Nerus Shift 2→3 ≥ 8 jam (-1 jam)"

    return 0.0, "", ""


def calculate_live_hours(raw_hours: float, day_type: str, deduction_hours: float = 0.0) -> float | None:
    raw = max(0.0, float(raw_hours or 0.0))
    deduction = max(0.0, float(deduction_hours or 0.0))
    effective = max(0.0, raw - deduction)
    if day_type == "day_off":
        return round(effective * 2.0, 6)
    if day_type == "workday":
        return round(min(effective, 1.0) * 1.5 + max(effective - 1.0, 0.0) * 2.0, 6)
    return None


def contract_period_for_company(company_key: str) -> dict | None:
    period = OVERTIME_COMPANY_CONTRACT_PERIODS.get(str(company_key or ""))
    if not period:
        return None
    return {
        "startDate": period["startDate"],
        "endDate": period["endDate"],
        "label": f"{period['startDate']} s.d. {period['endDate']}",
    }


def is_quota_eligible(work_date: str, company_key: str = "") -> bool:
    contract_period = contract_period_for_company(company_key)
    if not contract_period:
        return False
    normalized_date = str(work_date or "")
    return contract_period["startDate"] <= normalized_date <= contract_period["endDate"]


def quota_eligibility_label(work_date: str, company_key: str) -> str:
    if not contract_period_for_company(company_key):
        return "PT belum dipetakan"
    return "Termasuk periode kontrak" if is_quota_eligible(work_date, company_key) else "Di luar periode kontrak"


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip())


def company_from_contract(source_work_group: object) -> tuple[str, str]:
    """Map the source contract to its public company label without exposing the contract text."""
    normalized = _normalize_header(source_work_group).upper()
    normalized = re.sub(r"^NAMA KONTRAK\s+", "", normalized)
    for contract, company_key in OVERTIME_COMPANY_CONTRACTS:
        if normalized == contract or normalized.startswith(contract + " "):
            return company_key, OVERTIME_COMPANIES[company_key]
    return "unmapped", OVERTIME_COMPANIES["unmapped"]


def _decimal(value: object, label: str) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
    else:
        raw = str(value or "").strip().replace(" ", "")
        if not raw:
            raise ValueError(f"{label} kosong")
        if "," in raw and "." not in raw:
            raw = raw.replace(",", ".")
        try:
            number = float(raw)
        except ValueError as error:
            raise ValueError(f"{label} bukan angka yang valid") from error
    if not math.isfinite(number) or number < 0 or number > 48:
        raise ValueError(f"{label} harus berada di antara 0 dan 48 jam")
    return round(number, 6)


def _date_value(value: object, datemode: int) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and float(value) > 0:
        try:
            import xlrd

            return xlrd.xldate_as_datetime(float(value), datemode).date().isoformat()
        except (ValueError, OverflowError):
            pass
    raw = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw[:10], pattern).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Tanggal tidak valid: {raw or '-'}")


def _time_value(value: object, datemode: int) -> str:
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    if isinstance(value, (int, float)):
        try:
            import xlrd

            converted = xlrd.xldate_as_datetime(float(value), datemode)
            return converted.strftime("%H:%M:%S")
        except (ValueError, OverflowError):
            pass
    raw = str(value or "").strip()
    for pattern in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(raw, pattern).strftime("%H:%M:%S")
        except ValueError:
            continue
    raise ValueError(f"Jam tidak valid: {raw or '-'}")


def _datetime_value(value: object, datemode: int) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (int, float)):
        try:
            import xlrd

            return xlrd.xldate_as_datetime(float(value), datemode).isoformat(sep=" ", timespec="seconds")
        except (ValueError, OverflowError):
            pass
    raw = str(value).strip()
    for pattern in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, pattern).isoformat(sep=" ", timespec="seconds")
        except ValueError:
            continue
    return raw[:64]


def _clock_duration(start_time: str, end_time: str) -> float | None:
    try:
        start = datetime.strptime(start_time, "%H:%M:%S")
        end = datetime.strptime(end_time, "%H:%M:%S")
    except ValueError:
        return None
    if end < start:
        end += timedelta(days=1)
    return round((end - start).total_seconds() / 3600.0, 6)


def decode_overtime_file(file_name: str, file_data: str, max_bytes: int) -> bytes:
    if not str(file_name or "").strip().lower().endswith(".xls"):
        raise ValueError("File lembur wajib memakai format .xls")
    try:
        raw = base64.b64decode(str(file_data or ""), validate=True)
    except (ValueError, binascii.Error) as error:
        raise ValueError("Isi file .xls tidak valid") from error
    if not raw:
        raise ValueError("File .xls kosong")
    if len(raw) > max_bytes:
        raise ValueError(f"Ukuran file .xls melebihi batas {max_bytes // (1024 * 1024)} MB")
    if not raw.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
        raise ValueError("File bukan dokumen Excel .xls yang valid")
    return raw


def decode_overtime_schedule_file(file_name: str, file_data: str, max_bytes: int) -> bytes:
    if not str(file_name or "").strip().lower().endswith(".xlsx"):
        raise ValueError("Jadwal kerja wajib memakai format .xlsx")
    try:
        raw = base64.b64decode(str(file_data or ""), validate=True)
    except (ValueError, binascii.Error) as error:
        raise ValueError("Isi file jadwal kerja tidak valid") from error
    if not raw:
        raise ValueError("File jadwal kerja kosong")
    if len(raw) > max_bytes:
        raise ValueError(f"Ukuran file jadwal kerja melebihi batas {max_bytes // (1024 * 1024)} MB")
    if not raw.startswith(b"PK"):
        raise ValueError("File bukan dokumen Excel .xlsx yang valid")
    return raw


def _schedule_sheet_metadata(sheet_name: str) -> tuple[int, int] | None:
    normalized = re.sub(r"\s+", " ", str(sheet_name or "").strip().upper())
    base = re.fullmatch(r"(0(?:0[1-9]|1[0-2]))", normalized)
    if base:
        return int(base.group(1)), 0
    suffix_revision = re.fullmatch(r"(0(?:0[1-9]|1[0-2]))R(\d+)", normalized)
    if suffix_revision:
        return int(suffix_revision.group(1)), max(1, int(suffix_revision.group(2)))
    prefix_revision = re.fullmatch(r"R(\d+)(0(?:0[1-9]|1[0-2]))", normalized)
    if prefix_revision:
        return int(prefix_revision.group(2)), max(1, int(prefix_revision.group(1)))
    special_month = re.match(r"(0(?:0[1-9]|1[0-2]))\s+", normalized)
    if special_month:
        return int(special_month.group(1)), 1
    return None


def _schedule_year(values: list[object]) -> int:
    for value in values:
        match = re.search(r"\b(20\d{2})\b", str(value or ""))
        if match:
            return int(match.group(1))
    raise ValueError("Tahun tidak ditemukan pada header jadwal kerja")


def _schedule_day_type(shift_code: str) -> str:
    normalized = re.sub(r"\s+", " ", str(shift_code or "").strip().upper())
    if normalized in OVERTIME_SCHEDULE_OFF_CODES or normalized.startswith("LIBUR"):
        return "day_off"
    return "workday"


def parse_overtime_schedule_xlsx(file_bytes: bytes) -> tuple[list[dict], list[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ValueError("Parser .xlsx belum terpasang di server") from error
    try:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as error:
        raise ValueError("File jadwal kerja .xlsx rusak atau tidak dapat dibaca") from error
    try:
        selected_sheets = []
        candidates: dict[tuple[str, str], dict] = {}
        for sheet in workbook.worksheets:
            metadata = _schedule_sheet_metadata(sheet.title)
            if not metadata:
                continue
            month, revision_rank = metadata
            rows = sheet.iter_rows(values_only=True)
            header_rows = [next(rows, ()) for _ in range(8)]
            year = _schedule_year([value for row in header_rows[:4] for value in row])
            date_row = header_rows[7] if len(header_rows) >= 8 else ()
            day_columns = {
                index: int(value)
                for index, value in enumerate(date_row)
                if index >= 4 and isinstance(value, (int, float)) and 1 <= int(value) <= 31
            }
            if not day_columns:
                raise ValueError(f"Sheet {sheet.title} tidak memiliki baris tanggal yang valid")
            sheet_rows = 0
            for row in rows:
                employee_no = _clean_text(row[1] if len(row) > 1 else "")
                employee_name = _clean_text(row[2] if len(row) > 2 else "")
                if not employee_no or not employee_name:
                    continue
                for column, day in day_columns.items():
                    if column >= len(row):
                        continue
                    shift_code = _clean_text(row[column])
                    if not shift_code:
                        continue
                    try:
                        work_date = date(year, month, day).isoformat()
                    except ValueError:
                        continue
                    entry = {
                        "employeeNo": employee_no,
                        "employeeName": employee_name,
                        "workDate": work_date,
                        "shiftCode": shift_code,
                        "dayType": _schedule_day_type(shift_code),
                        "sourceSheet": sheet.title,
                        "revisionRank": revision_rank,
                    }
                    key = (employee_no, work_date)
                    existing = candidates.get(key)
                    if not existing or entry["revisionRank"] >= existing["revisionRank"]:
                        candidates[key] = entry
                    sheet_rows += 1
            selected_sheets.append({
                "name": sheet.title,
                "month": month,
                "revisionRank": revision_rank,
                "rows": sheet_rows,
            })
        if not selected_sheets:
            raise ValueError("Tidak ditemukan sheet jadwal bulanan seperti 001 atau 006R01")
        if not candidates:
            raise ValueError("Tidak ditemukan personel dan shift yang valid pada jadwal kerja")
        return list(candidates.values()), selected_sheets
    finally:
        workbook.close()


def preview_overtime_schedule_xlsx(payload: dict, max_bytes: int) -> dict:
    if not isinstance(payload, dict):
        raise ValueError("Payload upload harus berupa object")
    file_name = str(payload.get("fileName") or "").strip()
    file_bytes = decode_overtime_schedule_file(file_name, str(payload.get("fileData") or ""), max_bytes)
    entries, sheets = parse_overtime_schedule_xlsx(file_bytes)
    return {
        "preview": True,
        "fileName": file_name,
        "fileSha256": hashlib.sha256(file_bytes).hexdigest(),
        "entryCount": len(entries),
        "peopleCount": len({entry["employeeNo"] for entry in entries}),
        "sheets": sheets,
        "months": sorted({entry["workDate"][:7] for entry in entries}),
    }


def import_overtime_schedule_xlsx(
    connection_factory: Callable[[], sqlite3.Connection],
    payload: dict,
    user: dict,
    now_iso: str,
    max_bytes: int,
) -> dict:
    preview = preview_overtime_schedule_xlsx(payload, max_bytes)
    file_name = preview["fileName"]
    file_bytes = decode_overtime_schedule_file(file_name, str(payload.get("fileData") or ""), max_bytes)
    entries, sheets = parse_overtime_schedule_xlsx(file_bytes)
    batch_id = "ots-" + hashlib.sha256(f"{preview['fileSha256']}:{now_iso}".encode("utf-8")).hexdigest()[:24]
    with connection_factory() as connection:
        ensure_overtime_schema(connection)
        existing = connection.execute(
            "SELECT id, row_count FROM overtime_schedule_batches WHERE is_active = 1 AND file_sha256 = ?",
            (preview["fileSha256"],),
        ).fetchone()
        if existing:
            return {"noOp": True, "batchId": existing["id"], "imported": int(existing["row_count"] or 0), **preview}
        connection.execute("UPDATE overtime_schedule_batches SET is_active = 0, replaced_at = ? WHERE is_active = 1", (now_iso,))
        replaced = connection.execute("SELECT COUNT(*) AS total FROM overtime_schedule_entries").fetchone()
        connection.execute("DELETE FROM overtime_schedule_entries")
        connection.execute(
            "INSERT INTO overtime_schedule_batches (id, file_name, file_sha256, sheet_count, row_count, imported_by_user_id, imported_by_username, imported_at, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)",
            (batch_id, file_name[:240], preview["fileSha256"], len(sheets), len(entries), int(user["id"]), str(user["username"]), now_iso),
        )
        for entry in entries:
            connection.execute(
                "INSERT INTO overtime_schedule_entries (batch_id, employee_no, employee_name, work_date, shift_code, day_type, source_sheet, revision_rank) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (batch_id, entry["employeeNo"], entry["employeeName"], entry["workDate"], entry["shiftCode"], entry["dayType"], entry["sourceSheet"], entry["revisionRank"]),
            )
    return {"noOp": False, "batchId": batch_id, "imported": len(entries), "replaced": int(replaced["total"] or 0), **preview}


def parse_overtime_xls(file_bytes: bytes) -> list[dict]:
    try:
        import xlrd
    except ImportError as error:
        raise ValueError("Parser .xls belum terpasang di server") from error
    try:
        workbook = xlrd.open_workbook(file_contents=file_bytes, on_demand=True)
    except Exception as error:
        raise ValueError("File .xls rusak atau tidak dapat dibaca") from error
    try:
        sheet = workbook.sheet_by_name("data") if "data" in workbook.sheet_names() else workbook.sheet_by_index(0)
        if sheet.nrows < 2:
            raise ValueError("File .xls tidak memiliki baris transaksi")
        headers = [_normalize_header(sheet.cell_value(0, column)) for column in range(sheet.ncols)]
        missing = sorted(OVERTIME_REQUIRED_HEADERS.difference(headers))
        if missing:
            raise ValueError("Header file tidak lengkap: " + ", ".join(missing))
        index = {header: headers.index(header) for header in OVERTIME_REQUIRED_HEADERS}
        rows: list[dict] = []
        errors: list[str] = []
        for row_index in range(1, sheet.nrows):
            values = [sheet.cell_value(row_index, column) for column in range(sheet.ncols)]
            if not any(str(value or "").strip() for value in values):
                continue
            try:
                employee_no = _clean_text(values[index["no pegawai"]])
                employee_name = _clean_text(values[index["nama pegawai"]])
                if not employee_no or not employee_name:
                    raise ValueError("No Pegawai dan Nama Pegawai wajib diisi")
                work_date = _date_value(values[index["tanggal"]], workbook.datemode)
                start_time = _time_value(values[index["jam mulai"]], workbook.datemode)
                end_time = _time_value(values[index["jam selesai"]], workbook.datemode)
                raw_hours = _decimal(values[index["total"]], "Total")
                clock_hours = _clock_duration(start_time, end_time)
                source = {
                    "sourceNumber": _clean_text(values[index["no"]]),
                    "sourceStatus": _clean_text(values[index["status"]]),
                    "employeeNo": employee_no,
                    "employeeName": employee_name,
                    "workDate": work_date,
                    "period": work_date[:7],
                    "startTime": start_time,
                    "endTime": end_time,
                    "task": _clean_text(values[index["tugas"]]),
                    "rawHours": raw_hours,
                    "clockDurationHours": clock_hours,
                    "durationDelta": round(raw_hours - clock_hours, 6) if clock_hours is not None else None,
                    "unitName": _clean_text(values[index["unit kerja"]]),
                    "sourceWorkGroup": _clean_text(values[index["kelompok kerja"]]),
                    "approver": _clean_text(values[index["app by"]]),
                    "approvedAt": _datetime_value(values[index["app datetime"]], workbook.datemode),
                    "sourceRowNo": row_index + 1,
                }
                fingerprint_source = "\x1f".join(str(source[key]) for key in (
                    "sourceNumber", "sourceStatus", "employeeNo", "employeeName", "workDate",
                    "startTime", "endTime", "task", "rawHours", "unitName", "sourceWorkGroup",
                    "approver", "approvedAt",
                ))
                source["sourceFingerprint"] = hashlib.sha256(fingerprint_source.encode("utf-8")).hexdigest()
                source["id"] = "ot-" + source["sourceFingerprint"][:24]
                rows.append(source)
            except ValueError as error:
                errors.append(f"Baris {row_index + 1}: {error}")
        if errors:
            preview = "; ".join(errors[:8])
            suffix = f"; dan {len(errors) - 8} error lain" if len(errors) > 8 else ""
            raise ValueError(preview + suffix)
        if not rows:
            raise ValueError("Tidak ada transaksi valid dalam file .xls")
        return rows
    finally:
        workbook.release_resources()


def preview_overtime_xls(
    payload: dict,
    max_bytes: int,
    connection_factory: Callable[[], sqlite3.Connection] | None = None,
) -> dict:
    if not isinstance(payload, dict):
        raise ValueError("Payload upload harus berupa object")
    file_name = str(payload.get("fileName") or "").strip()
    file_bytes = decode_overtime_file(file_name, str(payload.get("fileData") or ""), max_bytes)
    rows = parse_overtime_xls(file_bytes)
    rows_by_period: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        rows_by_period[row["period"]].append(row)
    periods = sorted(rows_by_period)
    result = {
        "preview": True,
        "period": periods[0] if len(periods) == 1 else "",
        "periods": periods,
        "latestPeriod": periods[-1],
        "fileName": file_name,
        "fileSha256": hashlib.sha256(file_bytes).hexdigest(),
        "rowCount": len(rows),
        "peopleCount": len({row["employeeNo"] for row in rows}),
        "totalRawHours": round(sum(row["rawHours"] for row in rows), 2),
        "durationMismatchCount": sum(
            1
            for row in rows
            if row["durationDelta"] is not None and abs(row["durationDelta"]) >= 0.01
        ),
        "statuses": sorted({row["sourceStatus"] for row in rows if row["sourceStatus"]}),
        "dateFrom": min(row["workDate"] for row in rows),
        "dateTo": max(row["workDate"] for row in rows),
        "periodDetails": [
            {"period": period, "rowCount": len(rows_by_period[period])}
            for period in periods
        ],
    }
    if connection_factory is not None:
        with connection_factory() as connection:
            ensure_overtime_schema(connection)
            active_rows = connection.execute(
                "SELECT id, period, file_sha256, row_count FROM overtime_import_batches "
                "WHERE is_active = 1 AND period IN (" + ",".join("?" for _ in periods) + ")",
                periods,
            ).fetchall()
        active_by_period = {row["period"]: row for row in active_rows}
        details = []
        for detail in result["periodDetails"]:
            active = active_by_period.get(detail["period"])
            no_op = bool(active and active["file_sha256"] == result["fileSha256"])
            details.append({
                **detail,
                "willReplace": bool(active and not no_op),
                "noOp": no_op,
                "activeRowCount": int(active["row_count"] or 0) if active else 0,
            })
        result["periodDetails"] = details
        result["willReplace"] = any(detail["willReplace"] for detail in details)
        result["noOp"] = all(detail["noOp"] for detail in details)
        result["activeRowCount"] = sum(detail["activeRowCount"] for detail in details if detail["willReplace"])
    return result


def import_overtime_xls(
    connection_factory: Callable[[], sqlite3.Connection],
    payload: dict,
    user: dict,
    now_iso: str,
    max_bytes: int,
) -> dict:
    preview = preview_overtime_xls(payload, max_bytes)
    file_name = preview["fileName"]
    file_bytes = decode_overtime_file(file_name, str(payload.get("fileData") or ""), max_bytes)
    file_sha256 = preview["fileSha256"]
    rows = parse_overtime_xls(file_bytes)
    rows_by_period: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        rows_by_period[row["period"]].append(row)

    with connection_factory() as connection:
        ensure_overtime_schema(connection)
        active_rows = connection.execute(
            "SELECT id, period, file_sha256, row_count FROM overtime_import_batches "
            "WHERE is_active = 1 AND period IN (" + ",".join("?" for _ in rows_by_period) + ")",
            sorted(rows_by_period),
        ).fetchall()
        active_by_period = {row["period"]: row for row in active_rows}
        results = []
        for period in sorted(rows_by_period):
            period_rows = rows_by_period[period]
            existing = active_by_period.get(period)
            if existing and existing["file_sha256"] == file_sha256:
                results.append({"period": period, "noOp": True, "batchId": existing["id"], "imported": int(existing["row_count"] or 0), "replaced": 0})
                continue
            batch_id = "otb-" + hashlib.sha256(f"{period}:{file_sha256}:{now_iso}".encode("utf-8")).hexdigest()[:24]
            connection.execute(
                "UPDATE overtime_import_batches SET is_active = 0, replaced_at = ? WHERE period = ? AND is_active = 1",
                (now_iso, period),
            )
            replaced = connection.execute("SELECT COUNT(*) AS total FROM overtime_entries WHERE period = ?", (period,)).fetchone()
            replaced_count = int(replaced["total"] or 0)
            connection.execute("DELETE FROM overtime_entries WHERE period = ?", (period,))
            connection.execute(
                "INSERT INTO overtime_import_batches (id, period, file_name, file_sha256, row_count, imported_by_user_id, imported_by_username, imported_at, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)",
                (batch_id, period, file_name[:240], file_sha256, len(period_rows), int(user["id"]), str(user["username"]), now_iso),
            )
            for row in period_rows:
                connection.execute(
                    "INSERT INTO overtime_personnel (employee_no, employee_name, group_type, updated_by_user_id, updated_at) VALUES (?, ?, 'unassigned', ?, ?) ON CONFLICT(employee_no) DO UPDATE SET employee_name = excluded.employee_name",
                    (row["employeeNo"], row["employeeName"], int(user["id"]), now_iso),
                )
                connection.execute(
                    "INSERT INTO overtime_entries (id, period, batch_id, source_row_no, source_number, source_status, employee_no, employee_name, work_date, start_time, end_time, task, raw_hours, clock_duration_hours, duration_delta, unit_name, source_work_group, approver, approved_at, source_fingerprint, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (row["id"], period, batch_id, row["sourceRowNo"], row["sourceNumber"], row["sourceStatus"], row["employeeNo"], row["employeeName"], row["workDate"], row["startTime"], row["endTime"], row["task"], row["rawHours"], row["clockDurationHours"], row["durationDelta"], row["unitName"], row["sourceWorkGroup"], row["approver"], row["approvedAt"], row["sourceFingerprint"], now_iso),
                )
            results.append({"period": period, "noOp": False, "batchId": batch_id, "imported": len(period_rows), "replaced": replaced_count})
    return {
        "noOp": all(item["noOp"] for item in results),
        "period": preview["period"],
        "periods": [item["period"] for item in results],
        "latestPeriod": preview["latestPeriod"],
        "batchId": results[0]["batchId"] if len(results) == 1 else "",
        "imported": sum(item["imported"] for item in results if not item["noOp"]),
        "replaced": sum(item["replaced"] for item in results),
        "periodDetails": results,
        "fileName": file_name,
        "fileSha256": file_sha256,
    }


def _quota_status(hours: float) -> dict:
    if hours > OVERTIME_ANNUAL_QUOTA:
        return {"key": "exceeded", "label": "Kuota terlampaui"}
    if abs(hours - OVERTIME_ANNUAL_QUOTA) < 0.000001:
        return {"key": "exhausted", "label": "Kuota habis"}
    if hours >= OVERTIME_ANNUAL_QUOTA * OVERTIME_WARNING_RATIO:
        return {"key": "warning", "label": "Mendekati kuota"}
    return {"key": "normal", "label": "Normal"}


def _month_status(hours: float) -> dict:
    if hours > OVERTIME_MONTHLY_GUIDELINE:
        return {"key": "above", "label": "Di atas patokan"}
    if abs(hours - OVERTIME_MONTHLY_GUIDELINE) < 0.000001:
        return {"key": "at", "label": "Mencapai patokan"}
    return {"key": "below", "label": "Di bawah patokan"}


def _serialize_entry(row: sqlite3.Row) -> dict:
    day_type = str(row["day_type"] or "unknown")
    raw_hours = float(row["raw_hours"] or 0.0)
    company_key, company_name = company_from_contract(row["source_work_group"])
    contract_period = contract_period_for_company(company_key)
    schedule_shift_code = str(row["schedule_shift_code"] or "")
    start_time = str(row["start_time"] or "")
    end_time = str(row["end_time"] or "")

    deduction_hours, deduction_key, deduction_reason = calculate_overtime_deduction(
        company_key=company_key,
        day_type=day_type,
        raw_hours=raw_hours,
        start_time=start_time,
        end_time=end_time,
        schedule_shift_code=schedule_shift_code,
    )
    effective_raw_hours = max(0.0, raw_hours - deduction_hours)
    live_hours = calculate_live_hours(raw_hours, day_type, deduction_hours=deduction_hours)

    return {
        "id": row["id"],
        "period": row["period"],
        "status": row["source_status"],
        "employeeNo": row["employee_no"],
        "employeeName": row["employee_name"],
        "groupType": row["group_type"] or "unassigned",
        "companyKey": company_key,
        "companyName": company_name,
        "contractPeriod": contract_period,
        "workDate": row["work_date"],
        "startTime": row["start_time"],
        "endTime": row["end_time"],
        "task": row["task"],
        "rawHours": round(raw_hours, 2),
        "deductionHours": round(deduction_hours, 2),
        "effectiveRawHours": round(effective_raw_hours, 2),
        "deductionKey": deduction_key,
        "deductionReason": deduction_reason,
        "clockDurationHours": round(float(row["clock_duration_hours"]), 2) if row["clock_duration_hours"] is not None else None,
        "durationDelta": round(float(row["duration_delta"]), 2) if row["duration_delta"] is not None else None,
        "hasDurationMismatch": row["duration_delta"] is not None and abs(float(row["duration_delta"])) >= 0.01,
        "dayType": day_type,
        "classificationSource": row["classification_source"] or "unknown",
        "scheduleShiftCode": schedule_shift_code,
        "scheduleMatch": row["schedule_match"] or "",
        "liveHours": round(live_hours, 2) if live_hours is not None else None,
        "unitName": row["unit_name"],
        "approver": row["approver"],
        "approvedAt": row["approved_at"],
        "sourceRowNo": int(row["source_row_no"]),
        "quotaEligible": is_quota_eligible(row["work_date"], company_key),
        "quotaEligibilityLabel": quota_eligibility_label(row["work_date"], company_key),
    }


def _month_bounds(period: str) -> tuple[str, str]:
    if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", str(period or "")):
        raise ValueError("Parameter period wajib memakai format YYYY-MM")
    start = date.fromisoformat(period + "-01")
    end = (start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    return start.isoformat(), end.isoformat()


def _date_range(params: dict[str, str]) -> tuple[str, str]:
    start_date = str(params.get("startDate") or "").strip()
    end_date = str(params.get("endDate") or "").strip()
    if bool(start_date) != bool(end_date):
        raise ValueError("Tanggal awal dan akhir harus diisi bersama")
    if not start_date:
        return _month_bounds(str(params.get("period") or "").strip())
    try:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    except ValueError as error:
        raise ValueError("Rentang tanggal wajib memakai format YYYY-MM-DD") from error
    if start > end:
        raise ValueError("Tanggal awal tidak boleh setelah tanggal akhir")
    if (end - start).days > 730:
        raise ValueError("Rentang tanggal maksimal 730 hari")
    return start.isoformat(), end.isoformat()


def _load_quota_entries(connection: sqlite3.Connection, range_start: str, range_end: str) -> list[dict]:
    contract_start = min(item["startDate"] for item in OVERTIME_COMPANY_CONTRACT_PERIODS.values())
    contract_end = max(item["endDate"] for item in OVERTIME_COMPANY_CONTRACT_PERIODS.values())
    rows = connection.execute(
        """
        WITH active_schedule AS (
            SELECT employee_no, employee_name, work_date, shift_code, day_type
            FROM overtime_schedule_entries
            WHERE batch_id = (SELECT id FROM overtime_schedule_batches WHERE is_active = 1 LIMIT 1)
        ), schedule_by_name AS (
            SELECT lower(trim(employee_name)) AS employee_name_key, work_date,
                   max(shift_code) AS shift_code, max(day_type) AS day_type
            FROM active_schedule
            WHERE trim(employee_name) <> ''
            GROUP BY lower(trim(employee_name)), work_date
            HAVING count(*) = 1
        )
        SELECT e.*, COALESCE(p.group_type, 'unassigned') AS group_type,
               CASE WHEN d.employee_no IS NOT NULL THEN d.day_type
                    WHEN s.employee_no IS NOT NULL THEN s.day_type
                    WHEN sn.employee_name_key IS NOT NULL THEN sn.day_type
                    ELSE 'unknown' END AS day_type,
               CASE WHEN d.employee_no IS NOT NULL THEN 'manual'
                    WHEN s.employee_no IS NOT NULL OR sn.employee_name_key IS NOT NULL THEN 'schedule'
                    ELSE 'unknown' END AS classification_source,
               CASE WHEN s.employee_no IS NOT NULL THEN 'badge'
                    WHEN sn.employee_name_key IS NOT NULL THEN 'name'
                    ELSE '' END AS schedule_match,
               COALESCE(s.shift_code, sn.shift_code, '') AS schedule_shift_code
        FROM overtime_entries e
        LEFT JOIN overtime_personnel p ON p.employee_no = e.employee_no
        LEFT JOIN overtime_day_types d ON d.employee_no = e.employee_no AND d.work_date = e.work_date
        LEFT JOIN active_schedule s ON s.employee_no = e.employee_no AND s.work_date = e.work_date
        LEFT JOIN schedule_by_name sn ON sn.work_date = e.work_date AND sn.employee_name_key = lower(trim(e.employee_name))
            AND s.employee_no IS NULL
        WHERE e.work_date BETWEEN ? AND ? OR e.work_date BETWEEN ? AND ?
        ORDER BY e.work_date DESC, e.start_time DESC, e.source_row_no DESC
        """,
        (range_start, range_end, contract_start, contract_end),
    ).fetchall()
    return [_serialize_entry(row) for row in rows]


def _personnel_summary(entries: list[dict], range_start: str, range_end: str) -> list[dict]:
    buckets: dict[tuple[str, str], dict] = {}
    for item in entries:
        employee_no = item["employeeNo"]
        company_key = item["companyKey"]
        bucket = buckets.setdefault((employee_no, company_key), {
            "employeeNo": employee_no,
            "employeeName": item["employeeName"],
            "groupType": item["groupType"],
            "companyKey": company_key,
            "companyName": item["companyName"],
            "contractPeriod": item["contractPeriod"],
            "monthRawHours": 0.0,
            "monthLiveHours": 0.0,
            "contractLiveHours": 0.0,
            "unclassifiedCount": 0,
            "monthQuotaEligible": False,
        })
        if item["liveHours"] is None and item["quotaEligible"]:
            bucket["unclassifiedCount"] += 1
        elif item["quotaEligible"]:
            bucket["contractLiveHours"] += float(item["liveHours"])
        if range_start <= item["workDate"] <= range_end:
            bucket["monthRawHours"] += float(item["rawHours"])
            if item["liveHours"] is not None:
                bucket["monthLiveHours"] += float(item["liveHours"])
            bucket["monthQuotaEligible"] = bucket["monthQuotaEligible"] or item["quotaEligible"]
    result = []
    for bucket in buckets.values():
        for key in ("monthRawHours", "monthLiveHours", "contractLiveHours"):
            bucket[key] = round(bucket[key], 2)
        bucket["yearLiveHours"] = bucket["contractLiveHours"]
        if bucket["contractPeriod"]:
            bucket["annualRemainingHours"] = round(max(0.0, OVERTIME_ANNUAL_QUOTA - bucket["contractLiveHours"]), 2)
            bucket["annualOverHours"] = round(max(0.0, bucket["contractLiveHours"] - OVERTIME_ANNUAL_QUOTA), 2)
            bucket["annualUsagePercent"] = round(bucket["contractLiveHours"] / OVERTIME_ANNUAL_QUOTA * 100.0, 1)
            bucket["quotaStatus"] = _quota_status(bucket["contractLiveHours"])
            bucket["monthStatus"] = (
                _month_status(bucket["monthLiveHours"])
                if bucket["monthQuotaEligible"] else {"key": "outside_contract", "label": "Di luar periode kontrak"}
            )
        else:
            bucket["annualRemainingHours"] = None
            bucket["annualOverHours"] = 0.0
            bucket["annualUsagePercent"] = 0.0
            bucket["quotaStatus"] = {"key": "unmapped", "label": "PT belum dipetakan"}
            bucket["monthStatus"] = {"key": "unmapped", "label": "PT belum dipetakan"}
        bucket.pop("monthQuotaEligible")
        result.append(bucket)
    return sorted(result, key=lambda item: (-item["contractLiveHours"], item["employeeName"].lower(), item["companyName"]))


def _month_range(start_date: str, end_date: str) -> list[str]:
    cursor = date.fromisoformat(start_date).replace(day=1)
    end = date.fromisoformat(end_date).replace(day=1)
    months: list[str] = []
    while cursor <= end:
        months.append(cursor.strftime("%Y-%m"))
        cursor = (cursor.replace(day=28) + timedelta(days=4)).replace(day=1)
    return months


def _trend_periods(company_key: str) -> list[str]:
    contract_period = contract_period_for_company(company_key)
    if contract_period:
        return _month_range(contract_period["startDate"], contract_period["endDate"])
    starts = [item["startDate"] for item in OVERTIME_COMPANY_CONTRACT_PERIODS.values()]
    ends = [item["endDate"] for item in OVERTIME_COMPANY_CONTRACT_PERIODS.values()]
    return _month_range(min(starts), max(ends))


def list_overtime(
    connection_factory: Callable[[], sqlite3.Connection],
    params: dict[str, str],
) -> dict:
    range_start, range_end = _date_range(params)
    period = range_end[:7]
    try:
        page = max(1, int(params.get("page") or 1))
        limit = min(200, max(10, int(params.get("limit") or 50)))
    except ValueError as error:
        raise ValueError("Parameter pagination tidak valid") from error
    group_type = str(params.get("group") or "").strip().lower()
    day_type = str(params.get("dayType") or "").strip().lower()
    status = str(params.get("status") or "").strip().lower()
    company = str(params.get("company") or "").strip().lower()
    search = str(params.get("search") or "").strip().lower()
    if group_type and group_type not in OVERTIME_GROUPS:
        raise ValueError("Filter grup tidak valid")
    if day_type and day_type not in OVERTIME_DAY_TYPES:
        raise ValueError("Filter tipe hari tidak valid")
    if company and company not in OVERTIME_COMPANIES:
        raise ValueError("Filter PT tidak valid")

    with connection_factory() as connection:
        ensure_overtime_schema(connection)
        quota_entries = _load_quota_entries(connection, range_start, range_end)
        batches = connection.execute(
            "SELECT * FROM overtime_import_batches WHERE is_active = 1 AND period BETWEEN ? AND ? ORDER BY period, imported_at DESC",
            (range_start[:7], range_end[:7]),
        ).fetchall()
        schedule_batch = connection.execute(
            "SELECT * FROM overtime_schedule_batches WHERE is_active = 1 LIMIT 1"
        ).fetchone()
    selected_entries = [item for item in quota_entries if range_start <= item["workDate"] <= range_end]
    scoped_entries = [item for item in selected_entries if not company or item["companyKey"] == company]
    filtered = []
    for item in selected_entries:
        if group_type and item["groupType"] != group_type:
            continue
        if day_type and item["dayType"] != day_type:
            continue
        if status and item["status"].lower() != status:
            continue
        if company and item["companyKey"] != company:
            continue
        haystack = " ".join(str(item.get(key) or "") for key in (
            "employeeNo", "employeeName", "companyName", "task", "unitName", "status", "approver"
        )).lower()
        if search and search not in haystack:
            continue
        filtered.append(item)

    personnel = _personnel_summary(quota_entries, range_start, range_end)
    period_people = {(item["employeeNo"], item["companyKey"]) for item in scoped_entries}
    group_totals = {
        key: {"groupType": key, "people": 0, "rawHours": 0.0, "liveHours": 0.0, "unclassifiedCount": 0}
        for key in ("gangguan", "preventif", "unassigned")
    }
    group_people: dict[str, set[str]] = defaultdict(set)
    for item in scoped_entries:
        group_key = item["groupType"] if item["groupType"] in group_totals else "unassigned"
        group_people[group_key].add(item["employeeNo"])
        group_totals[group_key]["rawHours"] += item["rawHours"]
        if item["liveHours"] is None:
            group_totals[group_key]["unclassifiedCount"] += 1
        else:
            group_totals[group_key]["liveHours"] += item["liveHours"]
    for key, total in group_totals.items():
        total["people"] = len(group_people[key])
        total["rawHours"] = round(total["rawHours"], 2)
        total["liveHours"] = round(total["liveHours"], 2)

    trend = []
    for key in _trend_periods(company):
        values = [
            item for item in quota_entries
            if item["period"] == key and (not company or item["companyKey"] == company)
        ]
        trend.append({
            "period": key,
            "rawHours": round(sum(item["rawHours"] for item in values), 2),
            "liveHours": round(sum(item["liveHours"] or 0.0 for item in values), 2),
            "unclassifiedCount": sum(1 for item in values if item["liveHours"] is None),
        })

    total = len(filtered)
    start = (page - 1) * limit
    selected_personnel = [
        item for item in personnel
        if (item["employeeNo"], item["companyKey"]) in period_people
    ]
    quota_personnel = [item for item in selected_personnel if item["contractPeriod"]]
    annual_used = round(sum(item["contractLiveHours"] for item in quota_personnel), 2)
    annual_capacity = round(len(quota_personnel) * OVERTIME_ANNUAL_QUOTA, 2)
    return {
        "period": period,
        "range": {"startDate": range_start, "endDate": range_end},
        "items": filtered[start:start + limit],
        "pagination": {"page": page, "limit": limit, "total": total, "pages": max(1, math.ceil(total / limit))},
        "summary": {
            "totalRawHours": round(sum(item["rawHours"] for item in scoped_entries), 2),
            "totalLiveHours": round(sum(item["liveHours"] or 0.0 for item in scoped_entries), 2),
            "unclassifiedCount": sum(1 for item in scoped_entries if item["liveHours"] is None),
            "peopleCount": len(period_people),
            "annualUsedHours": annual_used,
            "annualCapacityHours": annual_capacity,
            "annualRemainingHours": round(max(0.0, annual_capacity - annual_used), 2),
            "monthlyGuidelinePerPerson": OVERTIME_MONTHLY_GUIDELINE,
            "annualQuotaPerPerson": OVERTIME_ANNUAL_QUOTA,
            "contractPeriods": {
                key: {"companyName": OVERTIME_COMPANIES[key], **contract_period_for_company(key)}
                for key in OVERTIME_COMPANY_CONTRACT_PERIODS
            },
            "groupTotals": list(group_totals.values()),
            "trend": trend,
            "ranking": selected_personnel,
        },
        "availableStatuses": sorted({item["status"] for item in selected_entries if item["status"]}),
        "availableCompanies": [
            {"key": key, "name": name}
            for key, name in OVERTIME_COMPANIES.items()
            if key != "unmapped" or any(item["companyKey"] == key for item in selected_entries)
        ],
        "activeBatch": dict(batches[0]) if len(batches) == 1 else None,
        "activeBatches": [dict(batch) for batch in batches],
        "activeScheduleBatch": dict(schedule_batch) if schedule_batch else None,
    }


def list_overtime_personnel(connection_factory: Callable[[], sqlite3.Connection], year: str, period: str | None = None) -> dict:
    if not re.fullmatch(r"\d{4}", str(year or "")):
        raise ValueError("Parameter year wajib memakai format YYYY")
    selected_period = period if period and re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", period) else year + "-01"
    range_start, range_end = _month_bounds(selected_period)
    with connection_factory() as connection:
        ensure_overtime_schema(connection)
        entries = _load_quota_entries(connection, range_start, range_end)
        rows = connection.execute(
            "SELECT employee_no, employee_name, group_type, updated_at FROM overtime_personnel ORDER BY lower(employee_name)"
        ).fetchall()
    items = _personnel_summary(entries, range_start, range_end)
    known_people = {item["employeeNo"] for item in items}
    for row in rows:
        if row["employee_no"] not in known_people:
            items.append({
                "employeeNo": row["employee_no"], "employeeName": row["employee_name"], "groupType": row["group_type"],
                "companyKey": "unmapped", "companyName": OVERTIME_COMPANIES["unmapped"], "contractPeriod": None,
                "monthRawHours": 0.0, "monthLiveHours": 0.0, "contractLiveHours": 0.0, "yearLiveHours": 0.0,
                "unclassifiedCount": 0, "annualRemainingHours": None, "annualOverHours": 0.0, "annualUsagePercent": 0.0,
                "quotaStatus": {"key": "unmapped", "label": "PT belum dipetakan"},
                "monthStatus": {"key": "unmapped", "label": "PT belum dipetakan"},
            })
    updated = {row["employee_no"]: row["updated_at"] for row in rows}
    for item in items:
        item["updatedAt"] = updated.get(item["employeeNo"], "")
    items.sort(key=lambda item: (item["employeeName"].lower(), item["companyName"]))
    return {"year": year, "period": selected_period, "items": items}


def update_overtime_personnel(
    connection_factory: Callable[[], sqlite3.Connection],
    employee_no: str,
    group_type: str,
    user_id: int,
    now_iso: str,
) -> dict:
    employee = str(employee_no or "").strip()
    group = str(group_type or "").strip().lower()
    if not employee:
        raise ValueError("No Pegawai wajib diisi")
    if group not in OVERTIME_GROUPS:
        raise ValueError("Grup personel tidak valid")
    with connection_factory() as connection:
        ensure_overtime_schema(connection)
        row = connection.execute("SELECT employee_name FROM overtime_personnel WHERE employee_no = ?", (employee,)).fetchone()
        if not row:
            raise ValueError("Personel tidak ditemukan")
        connection.execute(
            "UPDATE overtime_personnel SET group_type = ?, updated_by_user_id = ?, updated_at = ? WHERE employee_no = ?",
            (group, int(user_id), now_iso, employee),
        )
        name = row["employee_name"]
    return {"employeeNo": employee, "employeeName": name, "groupType": group, "updatedAt": now_iso}


def update_overtime_day_types(
    connection_factory: Callable[[], sqlite3.Connection],
    classifications: list,
    user_id: int,
    now_iso: str,
) -> dict:
    if not isinstance(classifications, list) or not classifications:
        raise ValueError("Daftar klasifikasi wajib diisi")
    if len(classifications) > 1000:
        raise ValueError("Maksimal 1000 klasifikasi per request")
    normalized: dict[tuple[str, str], str] = {}
    for item in classifications:
        if not isinstance(item, dict):
            raise ValueError("Format klasifikasi tidak valid")
        employee = str(item.get("employeeNo") or "").strip()
        work_date = str(item.get("workDate") or "").strip()
        day_type = str(item.get("dayType") or "").strip().lower()
        if not employee or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", work_date):
            raise ValueError("No Pegawai dan tanggal klasifikasi wajib valid")
        try:
            date.fromisoformat(work_date)
        except ValueError as error:
            raise ValueError(f"Tanggal klasifikasi tidak valid: {work_date}") from error
        if day_type not in OVERTIME_DAY_TYPES:
            raise ValueError("Tipe hari harus workday, day_off, atau unknown")
        normalized[(employee, work_date)] = day_type
    with connection_factory() as connection:
        ensure_overtime_schema(connection)
        for (employee, work_date), day_type in normalized.items():
            exists = connection.execute(
                "SELECT 1 FROM overtime_entries WHERE employee_no = ? AND work_date = ? LIMIT 1",
                (employee, work_date),
            ).fetchone()
            if not exists:
                raise ValueError(f"Transaksi {employee} pada {work_date} tidak ditemukan")
            if day_type == "unknown":
                connection.execute(
                    "DELETE FROM overtime_day_types WHERE employee_no = ? AND work_date = ?",
                    (employee, work_date),
                )
                continue
            connection.execute(
                """
                INSERT INTO overtime_day_types (employee_no, work_date, day_type, updated_by_user_id, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(employee_no, work_date) DO UPDATE SET
                    day_type = excluded.day_type,
                    updated_by_user_id = excluded.updated_by_user_id,
                    updated_at = excluded.updated_at
                """,
                (employee, work_date, day_type, int(user_id), now_iso),
            )
    return {"updated": len(normalized), "classifications": [
        {"employeeNo": employee, "workDate": work_date, "dayType": day_type}
        for (employee, work_date), day_type in normalized.items()
    ]}
